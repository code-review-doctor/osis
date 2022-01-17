##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2021 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import random

from typing import Union

from django.test import SimpleTestCase
from openpyxl import Workbook

from assessments.views.serializers.score_sheet_xls_import import ScoreSheetXLSImportSerializer, \
    ScoreSheetXLSImportSerializerError, ProgramManagerScoreSheetXLSImportSerializer
from unittest import mock
from assessments.export.score_sheet_xls import HEADER
from openpyxl.utils import get_column_letter


class ScoreSheetXLSImportSerializerTest(SimpleTestCase):
    def setUp(self) -> None:
        self.serializer_cls = ScoreSheetXLSImportSerializer
        self.workbook = Workbook()
        worksheet = self.workbook.active

        _add_row_to_worksheet(
            worksheet,
            13,
            annee_academique="2020-21",
            numero_session="2",
            code_cours="LEPL1509",
            noma="24641600",
            note="10",
            email="dummy@gmail.com",
        )

    @mock.patch(
        "assessments.views.serializers.score_sheet_xls_import.ScoreSheetXLSImportSerializer._check_headers_consistency"
    )
    def test_serialize_worksheet(self, mock_check_headers_consistency):
        score_sheet_serialized = self.serializer_cls(self.workbook.active).data

        self.assertEqual(score_sheet_serialized['annee_academique'], 2020)
        self.assertEqual(score_sheet_serialized['numero_session'], 2)

        self.assertEqual(len(score_sheet_serialized['notes_etudiants']), 1)
        self.assertEqual(score_sheet_serialized['notes_etudiants'][0]['code_unite_enseignement'],  "LEPL1509")
        self.assertEqual(score_sheet_serialized['notes_etudiants'][0]['note'], "10")
        self.assertEqual(score_sheet_serialized['notes_etudiants'][0]['email'], "dummy@gmail.com")
        self.assertEqual(score_sheet_serialized['notes_etudiants'][0]['noma'], "24641600")

    def test_assert_multiple_academic_year_found_raised_serialization_error(self):
        worksheet = self.workbook.active
        _add_row_to_worksheet(
            worksheet,
            14,
            annee_academique="2021-22",
            numero_session="2",
            code_cours="LEPL1509",
            noma="24641600",
            note="10",
            email="dummy@gmail.com",
        )

        with self.assertRaises(ScoreSheetXLSImportSerializerError):
            self.serializer_cls(self.workbook.active).data

    def test_assert_multiple_session_found_raised_serialization_error(self):
        worksheet = self.workbook.active
        _add_row_to_worksheet(
            worksheet,
            14,
            annee_academique="2020-21",
            numero_session="3",
            code_cours="LEPL1509",
            noma="24641600",
            note="10",
            email="dummy@gmail.com",
        )

        with self.assertRaises(ScoreSheetXLSImportSerializerError):
            self.serializer_cls(self.workbook.active).data

    @mock.patch(
        "assessments.views.serializers.score_sheet_xls_import.ScoreSheetXLSImportSerializer._check_headers_consistency")
    def test_assert_serialize_zero_as_value(self, mock_check_headers_consistency):
        worksheet = self.workbook.active
        _add_row_to_worksheet(
            worksheet,
            13,
            annee_academique="2020-21",
            numero_session="2",
            code_cours="LEPL1509",
            noma="24641600",
            note=0,
            email="dummy@gmail.com",
        )

        serialized_data = self.serializer_cls(self.workbook.active).data
        self.assertEqual(serialized_data['notes_etudiants'][0]['note'], '0')

    @mock.patch(
        "assessments.views.serializers.score_sheet_xls_import.ScoreSheetXLSImportSerializer._check_headers_consistency")
    def test_assert_convert_comma_to_dot(self, mock_check_headers_consistency):
        worksheet = self.workbook.active
        _add_row_to_worksheet(
            worksheet,
            13,
            annee_academique="2020-21",
            numero_session="2",
            code_cours="LEPL1509",
            noma="24641600",
            note="10,6",
            email="dummy@gmail.com",
        )
        serialized_data = self.serializer_cls(self.workbook.active).data
        self.assertEqual(serialized_data['notes_etudiants'][0]['note'], '10.6')

    def test_check_headers_consitency_no_headers(self):
        worksheet = self.workbook.active
        with self.assertRaises(ScoreSheetXLSImportSerializerError):
            self.serializer_cls._check_headers_consistency(worksheet)

    def test_check_headers_consitency_headers_ok(self):
        worksheet = self.workbook.active
        self._build_correct_headers(worksheet)
        self.serializer_cls._check_headers_consistency(worksheet)

    def test_check_headers_consitency_wrong_headers(self):
        worksheet = self.workbook.active
        self._build_correct_headers(worksheet)
        self.serializer_cls._check_headers_consistency(worksheet)

    @staticmethod
    def _build_correct_headers(worksheet):
        for count, header in enumerate(HEADER, start=1):
            worksheet[get_column_letter(count) + '1'].value = str(header)

    @staticmethod
    def _build_wrong_headers(worksheet):

        for count, header in enumerate(random.shuffle(HEADER), start=1):
            worksheet[get_column_letter(count) + '1'].value = str(header)


class ProgramManagerScoreSheetXLSImportSerializerTest(SimpleTestCase):
    def setUp(self) -> None:
        self.workbook = Workbook()
        self.serializer_cls = ProgramManagerScoreSheetXLSImportSerializer

    @mock.patch(
        "assessments.views.serializers.score_sheet_xls_import.ScoreSheetXLSImportSerializer._check_headers_consistency")
    def test_assert_A_letter_transformed_to_S_letter(self, mock_check_headers_consistency):
        worksheet = self.workbook.active

        _add_row_to_worksheet(
            worksheet,
            13,
            annee_academique="2020-21",
            numero_session="2",
            code_cours="LEPL1509",
            noma="24641600",
            note="A",
            email="dummy@gmail.com",
        )
        serialized_data = self.serializer_cls(self.workbook.active).data
        self.assertEqual(serialized_data['notes_etudiants'][0]['note'], 'S')


def _add_row_to_worksheet(
        worksheet,
        row_number,
        annee_academique: str,
        numero_session: str,
        code_cours: str,
        noma: str,
        note: Union[str, int],
        email: str
):
    worksheet['A' + str(row_number)].value = annee_academique
    worksheet['B' + str(row_number)].value = numero_session
    worksheet['C' + str(row_number)].value = code_cours
    worksheet['E' + str(row_number)].value = noma
    worksheet['H' + str(row_number)].value = note
    worksheet['G' + str(row_number)].value = email
