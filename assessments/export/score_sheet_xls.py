##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2021 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import decimal

from django.utils import timezone
from django.utils.translation import gettext_lazy as _, pgettext_lazy
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Color, PatternFill, Border, Side, Style
from openpyxl.styles.borders import BORDER_MEDIUM
from openpyxl.writer.excel import save_virtual_workbook

MAXIMAL_NUMBER_OF_DECIMALS = 1

HEADER_MANDATORY_PART = [
    _('Academic year'),
    _('Session'),
    _('Learning unit'),
    pgettext_lazy('encoding', 'Program'),
    pgettext_lazy('assessments', 'Registration number'),
    _('Name'),
    _('Email'),
    _('Score'),
    _('End date Prof'),
]

HEADER_OPTIONAL_PART = [
    _('Type of specific profile'),
    _('Extra time (33% generally)'),
    _('Large print'),
    _('Specific room of examination'),
    _('Other educational facilities'),
    _('Details other educational facilities'),
    _('Educational tutor'),
]

HEADER = HEADER_MANDATORY_PART + HEADER_OPTIONAL_PART


def build_xls(feuille_de_notes_serialized):
    workbook = Workbook()
    worksheet = workbook.active

    _build_document_info(worksheet, feuille_de_notes_serialized)
    _build_legends(worksheet, feuille_de_notes_serialized)
    _define_worksheet_column_size(worksheet)
    _build_headers(worksheet)
    _build_rows(worksheet, feuille_de_notes_serialized)
    worksheet.auto_filter.ref = "A11:P{}".format(str(worksheet.max_row))
    return save_virtual_workbook(workbook)


def _build_document_info(worksheet, feuille_de_notes_serialized):
    worksheet.append([
        feuille_de_notes_serialized['titre'],
        '',
        '',
        '',
        '',
        '',
        str('Contacts') if feuille_de_notes_serialized['contact_emails'] else '',
        feuille_de_notes_serialized['contact_emails'],
    ])
    worksheet.append([str('Session: %s' % feuille_de_notes_serialized['numero_session'])])
    worksheet.append([''])

    worksheet.cell(row=4, column=1).value = str(
        _("The data presented on this document correspond to the state of the system dated %(printing_date)s "
          "and are likely to evolve") % {'printing_date': timezone.now().strftime("%d/%m/%Y")}
    )
    worksheet.cell(row=4, column=1).font = Font(color=colors.RED)
    worksheet.cell(row=5, column=1).value = str(_("Students deliberated are not shown"))
    worksheet.cell(row=5, column=1).font = Font(color=colors.RED)
    worksheet.append([''])


def _build_legends(worksheet, feuille_de_notes_serialized):
    justification_other_values = "%s, %s" % (_('S=Unjustified Absence'), _('M=Justified Absence'))

    worksheet.append([
        str(_('Score')),
        str(_('Score legend: A=Absent, T=Cheating, {score} (0=Score of presence)').format(score="0 - 20")),
        str(''),
        str(''),
        str(''),
        str(''),
        str(_('Enrolled lately')),
    ])
    worksheet.append([
        str(''),
        str(_("Other values reserved to administration: %(justification_other_values)s ") % {
            'justification_other_values': justification_other_values}),
        str(''),
        str(''),
        str(''),
        str(''),
        str(_('Unsubscribed lately')),
    ])
    worksheet.append([
        str(''),
        str(_('Decimals authorized for this learning unit'))
        if feuille_de_notes_serialized['note_decimale_est_autorisee'] else
        str(_('Unauthorized decimal for this learning unit'))
    ])

    worksheet.cell(row=7, column=7).fill = PatternFill(patternType='solid', fgColor=Color('dff0d8'))
    worksheet.cell(row=8, column=7).fill = PatternFill(patternType='solid', fgColor=Color('f2dede'))
    worksheet.cell(row=9, column=2).font = Font(color=colors.RED)
    worksheet.append([str('')])


def _define_worksheet_column_size(worksheet):
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['C'].width = 18
    worksheet.column_dimensions['E'].width = 18
    worksheet.column_dimensions['F'].width = 35
    worksheet.column_dimensions['G'].width = 45
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 15
    worksheet.column_dimensions['K'].width = 20
    worksheet.column_dimensions['L'].width = 25
    worksheet.column_dimensions['M'].width = 15
    worksheet.column_dimensions['N'].width = 25
    worksheet.column_dimensions['O'].width = 25
    worksheet.column_dimensions['P'].width = 30


def _build_headers(worksheet):
    worksheet.append([str(header) for header in HEADER])


def _build_rows(worksheet, feuille_de_notes_serialized):
    current_row_number = 12

    for row in feuille_de_notes_serialized['rows']:
        note = _format_note(row['note'])
        worksheet.append([
            feuille_de_notes_serialized['annee_academique'],
            feuille_de_notes_serialized['numero_session'],
            feuille_de_notes_serialized['code_unite_enseignement'],
            row['nom_cohorte'],
            row['noma'],
            row['nom_complet'],
            row['email'],
            note,
            row['echeance_enseignant'],
            row['type_peps'],
            row['tiers_temps'],
            row['copie_adaptee'],
            row['local_specifique'],
            row['autre_amenagement'],
            row['details_autre_amenagement'],
            row['accompagnateur'],
        ])

        for column_number in range(1, 8):
            __set_non_editable_color(worksheet, column_number, current_row_number)
        if row['est_soumise']:
            __set_non_editable_color(worksheet, 8, current_row_number)
        __set_non_editable_color(worksheet, 9, current_row_number)

        if row['inscrit_tardivement']:
            __set_late_subscribe_row_color(worksheet, current_row_number)
        if row['desinscrit_tardivement']:
            __set_late_unsubscribe_row_color(worksheet, current_row_number)
        __set_border_on_first_peps_cell(worksheet, current_row_number)
        current_row_number += 1


def __set_non_editable_color(worksheet, column_number, row_number):
    pattern_fill_grey = PatternFill(patternType='solid', fgColor=Color('C1C1C1'))
    worksheet.cell(row=row_number, column=column_number).fill = pattern_fill_grey


def __set_late_subscribe_row_color(worksheet, row_number):
    pattern_fill_enrollment_state = PatternFill(patternType='solid', fgColor='dff0d8')
    for column_number in range(1, 10):
        worksheet.cell(row=row_number, column=column_number).fill = pattern_fill_enrollment_state


def __set_late_unsubscribe_row_color(worksheet, row_number):
    pattern_fill_enrollment_state = PatternFill(patternType='solid', fgColor='f2dede')
    for column_number in range(1, 10):
        worksheet.cell(row=row_number, column=column_number).fill = pattern_fill_enrollment_state


def __set_border_on_first_peps_cell(worksheet, row_number):
    first_peps_cell = worksheet["J{}".format(row_number)]
    medium_black_border = Border(
        left=Side(border_style=BORDER_MEDIUM, color=Color('FF000000')),
    )
    cell_style = first_peps_cell.style if first_peps_cell.has_style else Style()
    cell_style.border = medium_black_border

    first_peps_cell.style = cell_style


def _format_note(note: str) -> str:
    # TODO : Have to be deleted when examEnrollment score's field will be transform from decimal_places=2 to
    #  decimal_places=1
    try:
        number_of_decimal = decimal.Decimal(note).as_tuple().exponent * -1
        if number_of_decimal > MAXIMAL_NUMBER_OF_DECIMALS:
            return note[:-1]
    except decimal.DecimalException:
        # Note is a letter
        pass
    return note
