##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2021 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from collections import defaultdict
from decimal import Decimal
from typing import List, Dict, Any

from django.db.models import Case, When
from django.db.models.query import QuerySet
from django.utils.translation import gettext_lazy as _
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from base.business import learning_unit_year_with_context
from base.business.entity import build_entity_container_prefetch
from base.business.learning_unit import get_organization_from_learning_unit_year
from base.business.learning_unit_year_with_context import append_latest_entities, append_components, \
    get_learning_component_prefetch
from base.business.learning_units.comparison import get_partims_as_str
from base.business.proposal_xls import BLANK_VALUE, XLS_DESCRIPTION_COMPARISON, XLS_COMPARISON_FILENAME, \
    COMPARISON_PROPOSAL_TITLES, COMPARISON_WORKSHEET_TITLE, basic_titles, components_titles
from base.business.utils.convert import volume_format
from base.business.xls import get_name_or_username
from base.enums.component_detail import VOLUME_TOTAL, VOLUME_Q1, VOLUME_Q2, PLANNED_CLASSES, \
    VOLUME_REQUIREMENT_ENTITY, VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1, VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2, \
    VOLUME_TOTAL_REQUIREMENT_ENTITIES, REAL_CLASSES, VOLUME_GLOBAL
from base.models.academic_year import find_academic_year_by_id, AcademicYear
from base.models.campus import find_by_id as find_campus_by_id
from base.models.entity import find_by_id
from base.models.entity_version import EntityVersion
from base.models.enums import entity_container_year_link_type as entity_types, vacant_declaration_type, \
    attribution_procedure
from base.models.enums import learning_component_year_type
from base.models.enums.learning_component_year_type import LECTURING, PRACTICAL_EXERCISES
from base.models.enums.learning_container_year_types import LearningContainerYearType
from base.models.enums.learning_unit_year_periodicity import PERIODICITY_TYPES
from base.models.learning_unit_year import LearningUnitYear, get_by_id, LearningUnitYearQuerySet
from osis_common.document import xls_build
from reference.models.language import find_by_id as find_language_by_id

EMPTY_VALUE = ''
DATE_FORMAT = '%d-%m-%Y'
DATE_TIME_FORMAT = '%d-%m-%Y %H:%M'
DESC = "desc"
WORKSHEET_TITLE = _('Comparison of LUs')
XLS_FILENAME = _('learning_units_comparison')
XLS_DESCRIPTION = _("Comparison of learning units")

ACRONYM_COL_NUMBER = 0
ACADEMIC_COL_NUMBER = 1
CELLS_MODIFIED_NO_BORDER = 'modifications'
CELLS_TOP_BORDER = 'border_not_modified'
CELLS_STYLES = 'styles'
DATA = 'data'

REQUIREMENT_ENTITY_COL = "P"
ALLOCATION_ENTITY_COL = "Q"
ADDITIONAL_ENTITY_1_COL = "R"
ADDITIONAL_ENTITY_2_COL = "S"
STRIKETHROUGH_FONT = Font(strikethrough=True)
BOLD_FONT = Font(bold=True)
STYLE_MODIFIED_AND_ENTITY_INACTIVE = Font(color=Color('5CB85C'), strikethrough=True)
ENTITY_KEYS = ['requirement_entity', 'allocation_entity', 'additional_entity_1', 'additional_entity_2']
REQUIREMENT_ENTITY_COLUMN = 17
ALLOCATION_ENTITY_COLUMN = 18
ADDITIONAL_ENTITY_1_COLUMN = 19
ADDITIONAL_ENTITY_2_COLUMN = 20


def learning_unit_titles():
    return basic_titles() + components_titles()


def create_xls_comparison(user, learning_unit_years: QuerySet, filters, academic_yr_comparison: int):
    working_sheets_data = []
    cells_modified_with_green_font = []
    cells_with_top_border = []
    luys_for_2_years = None

    if learning_unit_years:
        luys_for_2_years = _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison,
                                                                       learning_unit_years)
        data = prepare_xls_content(luys_for_2_years)
        working_sheets_data = data.get('data')
        cells_modified_with_green_font = data.get(CELLS_MODIFIED_NO_BORDER)
        cells_with_top_border = data.get(CELLS_TOP_BORDER)
    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: learning_unit_titles(),
        xls_build.WS_TITLE: WORKSHEET_TITLE,
    }

    if cells_modified_with_green_font:
        parameters[xls_build.FONT_CELLS] = {xls_build.STYLE_MODIFIED: cells_modified_with_green_font}
    else:
        parameters[xls_build.FONT_CELLS] = {}
    if luys_for_2_years:
        parameters[xls_build.FONT_CELLS].update(
            _get_strikethrough_cells_on_entity(luys_for_2_years, cells_modified_with_green_font))
    if cells_with_top_border:
        parameters[xls_build.BORDER_CELLS] = {xls_build.BORDER_TOP: cells_with_top_border}
    parameters[xls_build.FONT_ROWS] = {BOLD_FONT: [0]}
    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison: int,
                                                learning_unit_years: QuerySet) -> List[LearningUnitYear]:
    l_units_order_to_preserve = [learning_unit_yr.learning_unit for learning_unit_yr in learning_unit_years]
    preserved = Case(*[When(learning_unit__pk=pk.id, then=pos) for pos, pk in enumerate(l_units_order_to_preserve)])

    learning_unit_years = LearningUnitYear.objects.filter(
        learning_unit__in=(_get_learning_units(learning_unit_years)),
        academic_year__year__in=(
            learning_unit_years[0].academic_year.year,
            academic_yr_comparison)
    ).select_related(
        'academic_year',
        'learning_container_year',
        'learning_container_year__academic_year'
    ).prefetch_related(
        get_learning_component_prefetch()
    ).prefetch_related(
        build_entity_container_prefetch(entity_types.ALLOCATION_ENTITY),
        build_entity_container_prefetch(entity_types.REQUIREMENT_ENTITY),
        build_entity_container_prefetch(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1),
        build_entity_container_prefetch(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2),
    ).order_by(preserved, 'academic_year__year')

    learning_unit_years = LearningUnitYearQuerySet\
        .annotate_entities_allocation_and_requirement_acronym(learning_unit_years)
    learning_unit_years = LearningUnitYearQuerySet.annotate_entities_status(learning_unit_years)
    learning_unit_years = LearningUnitYearQuerySet.annotate_additional_entities_status(learning_unit_years)

    [append_latest_entities(learning_unit) for learning_unit in learning_unit_years]
    [append_components(learning_unit) for learning_unit in learning_unit_years]
    return learning_unit_years


def _get_learning_units(learning_unit_years):
    return list(set([learning_unit_year.learning_unit for learning_unit_year in learning_unit_years]))


def prepare_xls_content(learning_unit_yrs):
    data = []
    learning_unit = None
    first_data = None
    modified_cells_no_border = []
    top_border = []
    for line_index, l_u_yr in enumerate(learning_unit_yrs, start=1):

        if learning_unit is None:
            learning_unit = l_u_yr.learning_unit
            new_line = True
        else:
            if learning_unit == l_u_yr.learning_unit:
                new_line = False
            else:
                learning_unit = l_u_yr.learning_unit
                new_line = True
        luy_data = extract_xls_data_from_learning_unit(l_u_yr, new_line, first_data)
        if new_line:
            first_data = luy_data
            top_border.extend(get_border_columns(line_index + 1))
        else:
            modified_cells_no_border.extend(
                _check_changes_other_than_code_and_year(first_data, luy_data, line_index + 1))
            first_data = None
        data.append(luy_data)

    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
    }


def extract_xls_data_from_learning_unit(learning_unit_yr, new_line, first_data):
    data = _get_data(learning_unit_yr, new_line, first_data)
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.LECTURING))
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.PRACTICAL_EXERCISES))
    return data


def translate_status(value):
    if value:
        return _('Active')
    else:
        return _('Inactive')


def _component_data(components, learning_component_yr_type):
    if components:
        for component in components:
            if component.type == learning_component_yr_type:
                return _get_volumes(component, components)
    return [EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE,
            EMPTY_VALUE, EMPTY_VALUE]


def _get_data(
        learning_unit_yr: LearningUnitYear,
        new_line: bool,
        first_data: List,
        partims: bool = True,
        proposal_comparison: bool = False
) -> List:
    organization = get_organization_from_learning_unit_year(learning_unit_yr)
    if proposal_comparison:
        academic_year = _format_academic_year(
            learning_unit_yr.academic_year.name,
            learning_unit_yr.learning_unit.end_year.name if learning_unit_yr.learning_unit.end_year else None
        )
    else:
        academic_year = learning_unit_yr.academic_year.name

    data = [
        _get_acronym(learning_unit_yr, new_line, first_data),
        academic_year,
        learning_unit_yr.learning_container_year.get_container_type_display()
        if learning_unit_yr.learning_container_year.container_type else BLANK_VALUE,
        translate_status(learning_unit_yr.status),
        learning_unit_yr.get_subtype_display() if learning_unit_yr.subtype else BLANK_VALUE,
        learning_unit_yr.get_internship_subtype_display() if learning_unit_yr.internship_subtype else BLANK_VALUE,
        volume_format(learning_unit_yr.credits) or BLANK_VALUE,
        learning_unit_yr.language.name or BLANK_VALUE,
        learning_unit_yr.get_periodicity_display() if learning_unit_yr.periodicity else BLANK_VALUE,
        learning_unit_yr.get_quadrimester_display() if learning_unit_yr.quadrimester else BLANK_VALUE,
        get_translation(learning_unit_yr.session),
        get_representing_string(learning_unit_yr.learning_container_year.common_title),
        get_representing_string(learning_unit_yr.specific_title),
        get_representing_string(learning_unit_yr.learning_container_year.common_title_english),
        get_representing_string(learning_unit_yr.specific_title_english),
        _get_entity_to_display(learning_unit_yr.ent_requirement_acronym),
        _get_entity_to_display(learning_unit_yr.ent_allocation_acronym),
        _get_entity_to_display(learning_unit_yr.additional_entity_1_acronym),
        _get_entity_to_display(learning_unit_yr.additional_entity_2_acronym),
        _('Yes') if learning_unit_yr.professional_integration else _('No'),
        organization.name if organization else BLANK_VALUE,
        learning_unit_yr.campus or BLANK_VALUE]
    if partims:
        data.append(get_partims_as_str(learning_unit_yr.get_partims_related()))
    data.extend(
        [
            get_representing_string(learning_unit_yr.faculty_remark),
            get_representing_string(learning_unit_yr.other_remark),
            get_representing_string(learning_unit_yr.other_remark_english),
            _('Yes') if learning_unit_yr.learning_container_year.team else _('No'),
            _('Yes') if learning_unit_yr.learning_container_year.is_vacant else _('No'),
            get_representing_string(learning_unit_yr.learning_container_year.get_type_declaration_vacant_display()),
            get_representing_string(learning_unit_yr.get_attribution_procedure_display()),
        ]
    )

    return data


def _get_acronym(learning_unit_yr, new_line, first_data):
    if first_data:
        acronym = EMPTY_VALUE
        if new_line:
            acronym = learning_unit_yr.acronym
        else:
            if learning_unit_yr.acronym != first_data[ACRONYM_COL_NUMBER]:
                acronym = learning_unit_yr.acronym
        return acronym
    else:
        return learning_unit_yr.acronym


def _get_volumes(component, components):
    volumes = components[component]
    return [
        volumes.get(VOLUME_Q1, EMPTY_VALUE),
        volumes.get(VOLUME_Q2, EMPTY_VALUE),
        volumes.get(VOLUME_TOTAL, EMPTY_VALUE),
        component.real_classes if component.real_classes else EMPTY_VALUE,
        component.planned_classes if component.planned_classes else EMPTY_VALUE,
        volumes.get(VOLUME_GLOBAL, '0'),
        volumes.get(VOLUME_REQUIREMENT_ENTITY, EMPTY_VALUE),
        volumes.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1, EMPTY_VALUE),
        volumes.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2, EMPTY_VALUE)
    ]


def get_translation(value):
    return str(_(value)) if value else BLANK_VALUE


def _get_entity_to_display(entity):
    return entity if entity else BLANK_VALUE


def _check_changes_other_than_code_and_year(first_data, second_data, line_index):
    modifications = []
    for col_index, obj in enumerate(first_data):
        if col_index == ACRONYM_COL_NUMBER and second_data[ACRONYM_COL_NUMBER] != EMPTY_VALUE:
            modifications.append('{}{}'.format(get_column_letter(col_index + 1), line_index))
        else:
            if obj != second_data[col_index] and col_index != ACADEMIC_COL_NUMBER:
                modifications.append('{}{}'.format(get_column_letter(col_index + 1), line_index))

    return modifications


def get_border_columns(line):
    style = []
    for col_index, obj in enumerate(learning_unit_titles(), start=1):
        style.append('{}{}'.format(get_column_letter(col_index), line))
    return style


def _get_component_data_by_type(component):
    if component:
        return [
            get_representing_string(component.get(VOLUME_Q1)),
            get_representing_string(component.get(VOLUME_Q2)),
            get_representing_string(component.get(VOLUME_TOTAL)),
            get_representing_string(component.get(REAL_CLASSES)),
            get_representing_string(component.get(PLANNED_CLASSES)),
            get_representing_string(component.get(VOLUME_TOTAL_REQUIREMENT_ENTITIES)),
            get_representing_string(component.get(VOLUME_REQUIREMENT_ENTITY)),
            get_representing_string(component.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1)),
            get_representing_string(component.get(VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2))
        ]
    else:
        return [BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE, BLANK_VALUE,
                BLANK_VALUE, BLANK_VALUE]


def prepare_xls_content_for_comparison(luy_with_proposals):
    line_index = 1
    data = []
    top_border = []
    modified_cells_no_border = []
    cells_font_styles = defaultdict(list)
    for luy_with_proposal in luy_with_proposals:
        top_border.extend(get_border_columns(line_index))
        data_proposal = _get_proposal_data(luy_with_proposal)
        data.append(data_proposal)

        proposal = luy_with_proposal.proposallearningunit
        initial_luy_data = proposal.initial_data
        entities_acronym_and_status_for_proposal = get_entities_acronym_and_status_for_proposal(luy_with_proposal)
        if initial_luy_data and initial_luy_data.get('learning_unit'):
            initial_data_info = _get_data_from_initial_data(luy_with_proposal, True)
            initial_data = initial_data_info['data']
            entities_acronym_and_status_for_initial_data = initial_data_info['entities_acronym_and_status']

            data.append(initial_data)
            cells_font_styles = _check_changes_in_proposals(
                initial_data,
                data_proposal,
                line_index + 1,
                entities_acronym_and_status_for_proposal,
                cells_font_styles
            )
            line_index += 2
            cells_font_styles = _check_entities_status_style_in_initial(
                cells_font_styles,
                entities_acronym_and_status_for_initial_data,
                line_index
            )
        else:
            cells_font_styles = _check_changes_in_proposals(
                None,
                data_proposal,
                line_index + 1,
                entities_acronym_and_status_for_proposal,
                cells_font_styles
            )
            line_index += 1
    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
        CELLS_STYLES: cells_font_styles
    }


def _get_data_from_initial_data(luy_with_proposal: LearningUnitYear, proposal_comparison=False) -> Dict[str, Any]:
    initial_data = luy_with_proposal.proposallearningunit.initial_data
    luy_initial = initial_data.get('learning_unit_year', {})
    lcy_initial = initial_data.get('learning_container_year', {})
    lu_initial = initial_data.get('learning_unit', {})

    if luy_initial.get('id'):
        learning_unit_yr = get_by_id(luy_initial.get('id'))
    else:
        learning_unit_yr = None

    entities_acronym_and_status = get_entities_acronym_and_status_for_initial_data(
        lcy_initial,
        luy_with_proposal.academic_year
    )
    campus = find_campus_by_id(luy_initial.get('campus'))

    organization = None
    if learning_unit_yr:
        organization = get_organization_from_learning_unit_year(learning_unit_yr)
    language = find_language_by_id(luy_initial.get('language'))

    if proposal_comparison:
        academic_year = _format_academic_year(learning_unit_yr.academic_year.name,
                                              find_academic_year_by_id(lu_initial.get('end_year'))
                                              if lu_initial.get('end_year') else None)
    else:
        academic_year = learning_unit_yr.academic_year.name

    data = [
        str(_('Initial data')),
        luy_initial.get('acronym', ''),
        academic_year,
        dict(LearningContainerYearType.choices())[lcy_initial.get('container_type')] if
        lcy_initial.get('container_type') else BLANK_VALUE,
        translate_status(luy_initial.get('status')),
        learning_unit_yr.get_subtype_display()
        if learning_unit_yr and learning_unit_yr.get_subtype_display() else BLANK_VALUE,
        get_translation(luy_initial.get('internship_subtype')),
        volume_format(Decimal(luy_initial['credits'])) if luy_initial.get('credits') else BLANK_VALUE,
        language.name if language else BLANK_VALUE,
        dict(PERIODICITY_TYPES)[luy_initial['periodicity']] if luy_initial.get('periodicity') else BLANK_VALUE,
        get_translation(luy_initial.get('quadrimester')),
        get_translation(luy_initial.get('session')),
        get_representing_string(lcy_initial.get('common_title')),
        get_representing_string(luy_initial.get('specific_title')),
        get_representing_string(lcy_initial.get('common_title_english')),
        get_representing_string(luy_initial.get('specific_title_english')),
        entities_acronym_and_status['requirement_entity']['acronym'],
        entities_acronym_and_status['allocation_entity']['acronym'],
        entities_acronym_and_status['additional_entity_1']['acronym'],
        entities_acronym_and_status['additional_entity_2']['acronym'],
        _('Yes') if luy_initial.get('professional_integration') else _('No'),
        organization.name if organization else BLANK_VALUE,
        campus if campus else BLANK_VALUE,
        get_representing_string(luy_initial.get('faculty_remark')),
        get_representing_string(luy_initial.get('other_remark')),
        get_representing_string(luy_initial.get('other_remark_english')),
        _('Yes') if lcy_initial.get('team') else _('No'),
        _('Yes') if lcy_initial.get('is_vacant') else _('No'),
        dict(vacant_declaration_type.DECLARATION_TYPE)[lcy_initial.get('type_declaration_vacant')] if lcy_initial.get(
            'type_declaration_vacant') else BLANK_VALUE,
        dict(attribution_procedure.ATTRIBUTION_PROCEDURES)[luy_initial.get('attribution_procedure')] if luy_initial.get(
            'attribution_procedure') else BLANK_VALUE,
    ]
    return {
        'data': _get_data_from_components_initial_data(data, initial_data),
        'entities_acronym_and_status': entities_acronym_and_status
    }


def get_representing_string(value):
    return value or BLANK_VALUE


def create_xls_proposal_comparison(user, lus_with_proposal, filters):
    lus_with_proposal = LearningUnitYearQuerySet.annotate_entities_status(lus_with_proposal)
    lus_with_proposal = LearningUnitYearQuerySet.annotate_additional_entities_status(lus_with_proposal)
    data = prepare_xls_content_for_comparison(lus_with_proposal)

    working_sheets_data = data.get('data')
    cells_with_top_border = data.get(CELLS_TOP_BORDER)

    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION_COMPARISON,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_COMPARISON_FILENAME,
        xls_build.HEADER_TITLES: COMPARISON_PROPOSAL_TITLES,
        xls_build.WS_TITLE: COMPARISON_WORKSHEET_TITLE,
    }

    parameters[xls_build.FONT_CELLS] = data.get(CELLS_STYLES, {})

    if cells_with_top_border:
        parameters[xls_build.BORDER_CELLS] = {xls_build.BORDER_BOTTOM: cells_with_top_border}
    parameters[xls_build.FONT_ROWS] = {BOLD_FONT: [0]}

    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_basic_components(learning_unit_yr):
    learning_unit_yr = find_learning_unit_yr_with_components_data(learning_unit_yr)
    components = []
    components_values = []
    for key, value in learning_unit_yr.components.items():
        components.append(key)
        components_values.append(value)

    practical_component = None
    lecturing_component = None
    for index, component in enumerate(components):
        if not practical_component and component.type == PRACTICAL_EXERCISES:
            practical_component = _build_component(component.real_classes, components_values, index)

        if not lecturing_component and component.type == LECTURING:
            lecturing_component = _build_component(component.real_classes, components_values, index)
    return {PRACTICAL_EXERCISES: practical_component, LECTURING: lecturing_component}


def _build_component(real_classes, components_values, index):
    a_component = components_values[index]
    a_component['REAL_CLASSES'] = real_classes
    return a_component


def _get_components_data(learning_unit_yr):
    components_data_dict = _get_basic_components(learning_unit_yr)
    return \
        _get_component_data_by_type(components_data_dict.get(LECTURING)) + \
        _get_component_data_by_type(components_data_dict.get(PRACTICAL_EXERCISES))


def _get_proposal_data(learning_unit_yr):
    data_proposal = [_('Proposal')] + _get_data(learning_unit_yr, False, None, False, True)
    data_proposal.extend(_get_components_data(learning_unit_yr))
    return data_proposal


def find_learning_unit_yr_with_components_data(learning_unit_yr):
    learning_unit_yrs = learning_unit_year_with_context.get_with_context(
        learning_container_year_id=learning_unit_yr.learning_container_year.id
    )
    if learning_unit_yrs:
        learning_unit_yr = next(luy for luy in learning_unit_yrs if luy.id == learning_unit_yr.id)

    return learning_unit_yr


def _get_data_from_components_initial_data(data_without_components, initial_data):
    data = data_without_components
    volumes = initial_data.get('volumes')
    if volumes:
        data = data + _get_component_data_by_type(volumes.get('PM'))
        data = data + _get_component_data_by_type(volumes.get('PP'))
    return data


def _format_academic_year(start_year, end_year) -> str:
    return "{}{}".format(start_year,
                         "   ({} {})".format(_('End').lower(), end_year if end_year else '-'))


def _get_strikethrough_cells_on_entity(luys: QuerySet, cells_modified_with_green_font: List) -> Dict:
    strikethrough_cells = defaultdict(list)
    strikethrough_cells.update({xls_build.STYLE_MODIFIED: cells_modified_with_green_font})

    for idx, luy in enumerate(list(luys), start=2):
        strikethrough_cells.update(
            _check_strike_cell_because_inactive_entity(
                cells_modified_with_green_font,
                luy.active_entity_requirement_version,
                strikethrough_cells,
                "{}{}".format(REQUIREMENT_ENTITY_COL, idx)
            )
        )
        strikethrough_cells.update(
            _check_strike_cell_because_inactive_entity(
                cells_modified_with_green_font,
                luy.active_entity_allocation_version,
                strikethrough_cells,
                "{}{}".format(ALLOCATION_ENTITY_COL, idx)
            )
        )
        strikethrough_cells.update(
            _check_strike_cell_because_inactive_entity(
                cells_modified_with_green_font,
                luy.active_additional_entity_1_version,
                strikethrough_cells,
                "{}{}".format(ADDITIONAL_ENTITY_1_COL, idx)
            )
        )
        strikethrough_cells.update(
            _check_strike_cell_because_inactive_entity(
                cells_modified_with_green_font,
                luy.active_additional_entity_2_version,
                strikethrough_cells,
                "{}{}".format(ADDITIONAL_ENTITY_2_COL, idx)
            )
        )
    return strikethrough_cells


def _check_strike_cell_because_inactive_entity(
        cells_modified_with_green_font: List,
        is_entity_active: bool,
        strikethrough_cells_param: Dict,
        cell_ref: str
) -> Dict:
    strike_and_modified_font = xls_build.STYLE_MODIFIED.copy()
    strike_and_modified_font.strikethrough = True
    strikethrough_cells = strikethrough_cells_param.copy()
    if not is_entity_active:
        if cell_ref in cells_modified_with_green_font:
            strikethrough_cells[strike_and_modified_font].append(cell_ref)
        else:
            strikethrough_cells[STRIKETHROUGH_FONT].append(cell_ref)
    return strikethrough_cells


def get_entities_acronym_and_status_for_initial_data(lcy_initial, academic_year):
    entities_acronym_and_status = {}
    for entity_key in ['requirement_entity', 'allocation_entity', 'additional_entity_1', 'additional_entity_2']:
        acronym = BLANK_VALUE
        entity = find_by_id(lcy_initial.get(entity_key))
        if entity:
            active_entity = EntityVersion.get_entity_if_active(entity, academic_year)
            if active_entity:
                acronym = active_entity.acronym
                active = True
            else:
                acronym = entity.most_recent_acronym
                active = False
        entities_acronym_and_status.update({entity_key: {'acronym': acronym, 'status': active}})
    return entities_acronym_and_status


def get_entities_acronym_and_status_for_proposal(luy: LearningUnitYear) -> Dict[str, Dict[str, Any]]:
    entities_acronym_and_status = {}

    for entity_key in ['requirement_entity', 'allocation_entity', 'additional_entity_1', 'additional_entity_2']:
        acronym = BLANK_VALUE
        entity = getattr(luy.learning_container_year, entity_key)
        if entity:
            active_entity = EntityVersion.get_entity_if_active(entity, luy.academic_year)
            if active_entity:
                acronym = active_entity.acronym
                active = True
            else:
                acronym = entity.most_recent_acronym
                active = False
        entities_acronym_and_status.update({entity_key: {'acronym': acronym, 'status': active}})
    return entities_acronym_and_status


def _check_entities_status_style_in_initial(
        styles_to_update: Dict[Font, str],
        entities_acronym_and_status: Dict,
        line_index: int
) -> Dict[Font, List[str]]:
    cells_font_styles = styles_to_update.copy()
    column = REQUIREMENT_ENTITY_COLUMN
    for key in ENTITY_KEYS:
        if not entities_acronym_and_status[key]['status']:
            cells_font_styles[STRIKETHROUGH_FONT].append("{}{}".format(get_column_letter(column), line_index))
        column += 1
    return cells_font_styles


def _check_changes_in_proposals(
        initial_data, proposal_data: list, line_index: int, entities_acronym_and_status: Dict, styles: Dict[Font, str]
) -> Dict[Font, List[str]]:

    if initial_data:
        for col_index, obj in enumerate(initial_data[2:], start=1):
            unactive_font = _get_unactive_font_status(col_index+2, entities_acronym_and_status)
            updated_font = _get_updated_font_status(col_index, obj, proposal_data)

            cell_ref = '{}{}'.format(get_column_letter(col_index+2), line_index)
            styles = _update_styles_for_cells(cell_ref, styles, unactive_font, updated_font)
    else:
        styles = _check_entities_status_style_in_initial(styles, entities_acronym_and_status, line_index)

    return styles


def _update_styles_for_cells(
        cell_ref: str,
        styles_to_update: Dict[Font, List[str]],
        unactive_font: bool,
        updated_font: bool
) -> Dict[Font, List[str]]:
    styles = styles_to_update.copy()
    if unactive_font and updated_font:
        styles[STYLE_MODIFIED_AND_ENTITY_INACTIVE].append(cell_ref)
    elif unactive_font:
        styles[STRIKETHROUGH_FONT].append(cell_ref)
    elif updated_font:
        styles[xls_build.STYLE_MODIFIED].append(cell_ref)
    return styles


def _get_updated_font_status(col_index, obj, proposal_data) -> bool:
    if str(obj) != str(proposal_data[col_index + 1]):
        return True
    return False


def _get_unactive_font_status(col_index: int, entities_acronym_and_status: Dict[str, Dict]) -> bool:
    if col_index == REQUIREMENT_ENTITY_COLUMN:
        if not entities_acronym_and_status['requirement_entity']['status']:
            return True
    if col_index == ALLOCATION_ENTITY_COLUMN:
        if not entities_acronym_and_status['allocation_entity']['status']:
            return True
    if col_index == ADDITIONAL_ENTITY_1_COLUMN:
        if not entities_acronym_and_status['additional_entity_1']['status']:
            return True
    if col_index == ADDITIONAL_ENTITY_2_COLUMN:
        if not entities_acronym_and_status['additional_entity_2']['status']:
            return True
    return False
